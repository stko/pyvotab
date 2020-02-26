import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pandas 
import numpy
from pyvotab import Pyvotab, PyvoSheet,PyvoStyles

class ptPlugin:

	def __init__(self):
		self.plugin_id='xls'

	def load(self, file_name):
		''' returns the excel file as dict of sheet_name:Panda_DataFrame
		'''
		res={}
		data_file = pandas.ExcelFile(file_name)			
		
		#für jedes sheet des dateipfads in pyvotab_list speichern
		for excelSheetname in data_file.sheet_names:
			res[excelSheetname] = pandas.read_excel (data_file, excelSheetname)
		return res


	def save(self, tables, input_file_name, file_name, options):
		print("plugin:",file_name)
		if not file_name.lower().endswith('.xlsx'):
			file_name = file_name+".xlsx"

		try:
			wb = load_workbook(input_file_name)
		except FileNotFoundError:
			wb = Workbook()
			wb.remove_sheet(wb.active)
			pass
		
		
		for i in range(len(tables)):
			sheet_name= re.sub('[^A-Za-z0-9._]+', '', tables[i].name)
			pt_table=tables[i].table
			template = tables[i].template

				
			for datasheet in wb:
				if(datasheet.title == sheet_name):
					wb.remove_sheet(datasheet) 
				
			if(template == None):	
				ws = wb.create_sheet(title=sheet_name)
			else:
				isSheetAvailable = False
				for datasheet in wb:
					if(datasheet.title == template):
						ws = wb.copy_worksheet(datasheet)
						ws.title = sheet_name
						isSheetAvailable = True

				if(not isSheetAvailable):
					ws = wb.create_sheet(title=sheet_name)
			try: # try if a style is set
				ws.sheet_properties.tabColor = tables[i].style['xls']
			except:
				pass
			try:
				for row in range(pt_table.ySize):
					for col in range(pt_table.xSize):
						try:
							cell_content=pt_table[col][row]
							if cell_content:
								this_cell=ws.cell(column=col+1, row=row+1)
								this_cell.value="{0}".format(str(cell_content["value"]))
								try:
									this_style=cell_content["style"][self.plugin_id]
									if this_style:
										#this_cell.fill = PatternFill("solid", fgColor=this_style)
										this_cell.fill = PatternFill("lightDown", fgColor=this_style)
								except:
									pass
								if cell_content["size"]>1:
									if cell_content["xDir"]:
										ws.merge_cells(start_row=row+1, start_column=col+1, end_row=row+1+cell_content["size"]-1, end_column=col+1)
									else:
										ws.merge_cells(start_row=row+1, start_column=col+1, end_row=row+1, end_column=col+1+cell_content["size"]-1)
									this_cell.alignment = Alignment(horizontal="center", vertical="center")
						except KeyError:
							pass
			except AttributeError:

				for row in range(0, len(pt_table)):
					for col in range(len(pt_table[row])):#pt_table[i][j]
						this_cell=ws.cell(row=row+1,column=col+1)
						this_cell.value=str(pt_table[row][col])
						#this_cell.fill = PatternFill("solid", fgColor="white")
					
			if(tables[i].name == "pyvotab"):
				ws.cell(row=1, column=1).value = "made with pyvotab"
				ws.cell(row=1, column=1).hyperlink = "https://github.com/stko/pyvotab"
				ws.cell(row=1, column=1).style = "Hyperlink"	
			
		wb.save(filename = file_name)
		
	def convertToExcelString(self, text):
		text = re.sub('[^A-Za-z0-9._]+', '', text)
		return text[:30]
		
	def containSheet(self, sheetold, sheetnew):
		sheetold = self.convertToExcelString(sheetold)
		sheetnew = self.convertToExcelString(sheetnew)

		if(sheetold.upper() == sheetnew.upper()):
			return True
		else:
			return False
		
	def calculate_pyvotab_files(self, file_name_list, local_layout_string_list, pyvot_style):
		'''
		Calculates the new pyvotabs out of a given set of excel input files

		Parameters
		----------
		
		file_name_list: List of file names

			List of pyvotab excel files which should be overlayed. The last file in the list contains the pyvotab layout instructions to be used
			and is used as template for the final output file
		
		local_layout_string_list: List of layout string
		
				Additional layout strings (optional), which are calculated on top of the ones which are already in the excel table

		pyvot_style : PyvotStyle

			contains the different format styles to highlight some different changes in the result tables
		'''

		if not file_name_list:
			return {} # no input, so no output
		
		layout_list_sheet = []
		view_sheet_list=[]
		pyvotab_sheet_list =[]

		# für jedes Sheet in der selektierten Datei
		for excelSheetname , data_file_element in self.load(file_name_list[-1]).items(): # open the last file in list
			#für jedes sheet des dateipfads in pyvotab_list speichern
			this_sheet = data_file_element.as_matrix()
			header = []
			for col in data_file_element.columns:
				header.append(col)
			#eingabebefehl in pyvotab speichern
			if excelSheetname == "pyvotab":
				layout_list_sheet = this_sheet
				if local_layout_string_list:
					layout_list_sheet = numpy.append(this_sheet, [local_layout_string_list], axis=0)	
			this_sheet = numpy.append([header], this_sheet, axis=0)
			pyvo = PyvoSheet(excelSheetname, this_sheet, None, None)
			view_sheet_list.append(pyvo)
			if excelSheetname == "pyvotab" or excelSheetname.startswith("pt."):
				pyvotab_sheet_list.append(pyvo)
		#neues pyvotabsheet erstellen
		# https://stackoverflow.com/a/11295857
		if layout_list_sheet.size==0:
			header = [["layout"]]
			layout_list_sheet[[]]
			if local_layout_string_list:
				layout_list_sheet = [[local_layout_string_list]]
			layout_list_sheet = numpy.append([header], layout_list_sheet, axis=0)
			pyvo = PyvoSheet("pyvotab", layout_list_sheet, None, None)
			view_sheet_list.append(pyvo)
			pyvotab_sheet_list.append(pyvo)



		'''
		#Unterscheidung zwischen Abarbeiten aller Befehle vom Pyvotab und nur das Abarbeiten der Eingabe
		for pyvo in view_list:
			if pyvo.name == "pyvotab":
				if local_layout_string_list:
					layout_list_sheet = [["layout"],[local_layout_string_list]]
				else:
					layout_list_sheet = pyvo.table
				break
		'''
		#Alle Befehle in save_list werden abgearbeitet	
		for actual_layout_string in map(lambda cell : cell[0], layout_list_sheet):
			if type(actual_layout_string) is str and not actual_layout_string=="":
				pyvotab = Pyvotab(pyvot_style, actual_layout_string, debug=False)	
				pt_name=pyvotab.get_url_parameter(pyvotab.layout,"source","pt.1")
				last_file_name=file_name_list[-1]
				for actual_file_name in file_name_list:
					try: 
						data_file_element = self.load(actual_file_name)[pt_name]
						header =list(data_file_element.columns)
						save_list=data_file_element.as_matrix()
						save_list= numpy.append([header] , save_list,axis =0)
						pyvotab.InsertTable( save_list.tolist(), last_file_name==actual_file_name, None)
					except:
						print("Error: Can not load pivotab Sheet {0}",pt_name)
				printDict = pyvotab.getPrintDict() # add result to global result table				
				pyvotab_sheet_list +=	printDict	
				view_sheet_list += printDict
		return pyvotab_sheet_list , view_sheet_list
